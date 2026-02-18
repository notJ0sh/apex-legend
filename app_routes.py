#      -----      {{{     IMPORTS     }}}      -----      #

from flask import render_template, redirect, url_for, request
from flask_login import current_user
from auth_user_routes import register_auth_routes
from database_helpers import get_database, USER_DATABASE, get_files_by_department, FILES_DATABASE
from models import File
import os
from datetime import datetime
from werkzeug.utils import secure_filename

#      -----      {{{     ROUTES (MAIN EVENTS)     }}}      -----      #

def get_file_icon(file_type):
    """Maps file extensions to emojis for the UI."""
    icon_map = {
        'pdf': '📄',
        'word': '📄',
        'doc': '📝',
        'docx': '📝',
        'ppt': '📊',
        'pptx': '📊',
        'xls': '📈',
        'xlsx': '📈',
        'jpg': '🖼️',
        'jpeg': '🖼️',
        'png': '🖼️',
        'zip': '📦',
        'txt': '📃',
        'csv': '📋'
    }
    return icon_map.get(file_type.lower(), '📎')

def format_datetime(datetime_str):
    """Format datetime string to DD-MM-YYYY HH:MM:SS (24h format)."""
    if not datetime_str:
        return 'N/A'
    
    try:
        from datetime import datetime
        dt = datetime.fromisoformat(datetime_str.replace('Z', '+00:00'))
        return dt.strftime('%d-%m-%Y %H:%M:%S')
    except (ValueError, AttributeError):
        return str(datetime_str)[:19]

def register_routes(app):
    """Register all routes with the Flask app."""

    register_auth_routes(app)

    #Settings page
    @app.route('/settings')
    def settings():
        # Fetch the current user's email and phone from database
        db = get_database(USER_DATABASE)
        user_data = db.execute(
            'SELECT email, phone_number FROM users WHERE id = ?',
            (current_user.id,)
        ).fetchone()
        
        email = user_data[0] if user_data else None
        phone_number = user_data[1] if user_data else None
        
        return render_template('settings.html', 
                             user_email=email, 
                             user_phone=phone_number)
    
    #Feedback submission -> saves it to Excel file for devs only
    @app.route('/submit-feedback', methods=['POST'])
    def submit_feedback():
        """Handle feedback form submission - saves to Excel file."""
        if not current_user.is_authenticated:
            return "Unauthorized", 401
        
        try:
            from openpyxl import Workbook, load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            # Get form data
            feedback_type = request.form.get('feedback_type', '').strip()
            subject = request.form.get('subject', '').strip()
            message = request.form.get('message', '').strip()
            include_contact = request.form.get('include_contact', 'false')

            # FIX: properly parse checkbox - 'true' string from JS
            include_contact_bool = include_contact == 'true'

            # Validation
            if not feedback_type or not subject or not message:
                return "All fields are required", 400

            # Get user contact info if they opted in
            user_email = 'N/A'
            user_phone = 'N/A'

            if include_contact_bool:
                db = get_database(USER_DATABASE)
                user_data = db.execute(
                    'SELECT email, phone_number FROM users WHERE id = ?',
                    (current_user.id,)
                ).fetchone()
                if user_data:
                    user_email = user_data[0] if user_data[0] else 'Not set'
                    user_phone = user_data[1] if user_data[1] else 'Not set'

            # FIX: save to .xlsx not .csv
            feedback_file = 'CLIENT_FEEDBACK.xlsx'

            if os.path.exists(feedback_file):
                wb = load_workbook(feedback_file)
                ws = wb.active
            else:
                # Create new file with styled headers
                wb = Workbook()
                ws = wb.active
                ws.title = "Client Feedback"

                headers = ['Timestamp', 'User ID', 'Username', 'Department', 'Role',
                           'Feedback Type', 'Subject', 'Message', 'Wants Contact', 'Email', 'Phone']
                ws.append(headers)

                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF')
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 10
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 10
                ws.column_dimensions['F'].width = 22
                ws.column_dimensions['G'].width = 30
                ws.column_dimensions['H'].width = 50
                ws.column_dimensions['I'].width = 15
                ws.column_dimensions['J'].width = 25
                ws.column_dimensions['K'].width = 15

            ws.append([
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                current_user.id,
                current_user.username,
                current_user.department if hasattr(current_user, 'department') and current_user.department else 'N/A',
                current_user.role,
                feedback_type,
                subject,
                message,
                'Yes' if include_contact_bool else 'No',
                user_email,
                user_phone
            ])

            wb.save(feedback_file)
            print(f"✅ Feedback saved to {feedback_file} | Wants contact: {include_contact_bool}")
            return "Feedback submitted successfully", 200

        except Exception as e:
            import traceback
            print(f"❌ Error submitting feedback: {e}")
            print(traceback.format_exc())
            return f"Failed to submit feedback: {str(e)}", 500
        
    #View feedback via webapp
    @app.route('/view-feedback-live')
    def view_feedback_live():
        """View feedback in browser - reads from Excel file."""
        if not current_user.is_authenticated:
            return "Unauthorized", 401
        
        try:
            from openpyxl import load_workbook

            # FIX: match actual filename on disk
            feedback_file = 'CLIENT_FEEDBACK.xlsx'

            if not os.path.exists(feedback_file):
                return render_template('feedback-viewer.html', feedback_data=[], message="No feedback submitted yet.")

            # Read-only mode so it doesn't lock the file
            wb = load_workbook(feedback_file, read_only=True)
            ws = wb.active

            feedback_data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # skip empty rows
                    feedback_data.append(row)

            wb.close()

            # Newest first
            feedback_data.reverse()

            # FIX: correct template name matches actual file 'feedback-viewer.html'
            return render_template('feedback-viewer.html', feedback_data=feedback_data)

        except Exception as e:
            import traceback
            print(f"❌ Error viewing feedback: {e}")
            print(traceback.format_exc())
            return f"Error loading feedback: {str(e)}", 500

    
    #Validation for updating user's details
    @app.route('/update-profile', methods=['POST'])
    def update_profile():
        email = request.form.get('email', '').strip()
        phone_number = request.form.get('phone_number', '').strip()
        
        if email and '@' not in email:
            return "Invalid email format", 400
        
        if phone_number and (not phone_number.isdigit() or len(phone_number) != 8):
            return "Phone number must be exactly 8 digits", 400
        
        db = get_database(USER_DATABASE)
        try:
            db.execute(
                'UPDATE users SET email = ?, phone_number = ? WHERE id = ?',
                (email if email else None, phone_number if phone_number else None, current_user.id)
            )
            db.commit()
            return redirect(url_for('settings'))
        except Exception as e:
            print(f"Error updating profile: {e}")
            return "Failed to update profile", 500
        
    @app.route('/files')
    def files():
        db = get_database(FILES_DATABASE)

        # --- ADDITIONAL FEATURE: STATS SUMMARY ---
        # This aggregates the count of files by source (Discord vs Web)
        stats_query = db.execute('SELECT source, COUNT(*) as count FROM files GROUP BY source').fetchall()
        stats_summary = {row['source']: row['count'] for row in stats_query}

        # Get all unique departments for filter dropdown
        departments_rows = db.execute(
            'SELECT DISTINCT department FROM files ORDER BY department'
        ).fetchall()
        departments = [dept[0] for dept in departments_rows if dept[0]]
        
        # Get all unique file types for filter dropdown
        file_types_query = db.execute(
            'SELECT DISTINCT file_type FROM files ORDER BY file_type'
        ).fetchall()
        file_types = [ft[0] for ft in file_types_query if ft[0]]
        
        # Get selected filters from query parameters
        selected_dept = request.args.get('department', 'all')
        selected_file_type = request.args.get('file_type', 'all')
        search_query = request.args.get('search', '').strip()

        query = 'SELECT * FROM files'
        conditions = []
        params = []

        if selected_dept and selected_dept != 'all':
            conditions.append('department = ?')
            params.append(selected_dept)

        if selected_file_type and selected_file_type != 'all':
            conditions.append('LOWER(file_type) = LOWER(?)')
            params.append(selected_file_type)

        if search_query:
            conditions.append('file_name LIKE ?')
            params.append(f'%{search_query}%')
        
        if conditions:
            query += ' WHERE ' + ' AND '.join(conditions)
        
        query += ' ORDER BY time_stamp DESC'

        cursor = db.execute(query, tuple(params)).fetchall()
        files_list = [File.from_row(row) for row in cursor]
    
        return render_template('files.html',
                            files=files_list,
                            stats_summary=stats_summary, # Added for reporting feature
                            departments=departments,
                            file_types=file_types,
                            selected_dept=selected_dept,
                            selected_file_type=selected_file_type,
                            search_query=search_query,
                            get_file_icon=get_file_icon,
                            format_datetime=format_datetime)

    @app.route('/download/<filename>')
    def download_file(filename):
        from database_helpers import get_file_download
        return get_file_download(filename)
    
    @app.route('/edit-file/<int:file_id>')
    def edit_file(file_id):
        if current_user.role != 'admin':
            return "Unauthorized - Admin access required", 403
        
        db = get_database(FILES_DATABASE)
        file_data = db.execute('SELECT * FROM files WHERE id = ?', (file_id,)).fetchone()
        
        if not file_data:
            return "File not found", 404
        
        file = File.from_row(file_data)
        
        departments_rows = db.execute('SELECT DISTINCT department FROM files ORDER BY department').fetchall()
        departments = [dept[0] for dept in departments_rows if dept[0]]
        
        return render_template('edit-file.html', file=file, departments=departments)

    @app.route('/update-file/<int:file_id>', methods=['POST'])
    def update_file(file_id):
        if current_user.role != 'admin':
            return "Unauthorized - Admin access required", 403
        
        file_name = request.form.get('file_name', '').strip()
        department = request.form.get('department', '').strip()
        project = request.form.get('project', '').strip()
        
        if not file_name or not project:
            return "File name and project cannot be empty", 400
        
        if file_name.isdigit() or project.isdigit():
            return "Fields cannot be just numbers", 400
        
        db = get_database(FILES_DATABASE)
        current_file = db.execute('SELECT * FROM files WHERE id = ?', (file_id,)).fetchone()
        
        if not current_file:
            return "File not found", 404
        
        try:
            db.execute(
                'UPDATE files SET file_name = ?, department = ?, project = ? WHERE id = ?',
                (file_name, department, project, file_id)
            )
            db.commit()
            
            if file_name != current_file['file_name']:
                old_path = os.path.join('downloads', current_file['file_name'])
                new_path = os.path.join('downloads', secure_filename(file_name))
                if os.path.exists(old_path):
                    os.rename(old_path, new_path)
                    db.execute('UPDATE files SET file_path = ? WHERE id = ?', (new_path, file_id))
                    db.commit()
            
            return redirect(url_for('files'))
        except Exception as e:
            print(f"Error updating file: {e}")
            return "Failed to update file", 500

    @app.route('/delete-file/<int:file_id>', methods=['POST'])
    def delete_file(file_id):
        if current_user.role != 'admin':
            return "Unauthorized - Admin access required", 403
        
        db = get_database(FILES_DATABASE)
        file_data = db.execute('SELECT * FROM files WHERE id = ?', (file_id,)).fetchone()
        
        if not file_data:
            return "File not found", 404
        
        try:
            db.execute('DELETE FROM files WHERE id = ?', (file_id,))
            db.commit()
            
            file_path = os.path.join('downloads', file_data['file_name'])
            if os.path.exists(file_path):
                os.remove(file_path)
            
            return redirect(url_for('files'))
        except Exception as e:
            print(f"Error deleting file: {e}")
            return "Failed to delete file", 500
